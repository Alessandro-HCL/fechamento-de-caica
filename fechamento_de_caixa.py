# import streamlit as st
# import pandas as pd
# from datetime import datetime
# import yagmail

# # Inicializa o estado da sessão para armazenar os dados do fechamento
# if 'fechamento' not in st.session_state:
#     st.session_state.fechamento = {}

# st.title("💳 Fechamento de Caixa - Villa Sonali")

# # Entradas do colaborador
# valor_pix = st.number_input("💵 Valor recebido em PIX:", min_value=0.0, step=0.01)
# valor_dinheiro = st.number_input("💵 Valor recebido em Dinheiro:", min_value=0.0, step=0.01)
# valor_cartao = st.number_input("💳 Valor recebido em Cartão:", min_value=0.0, step=0.01)
# valor_pendura = st.number_input("📞 Valor em Pendura:", min_value=0.0, step=0.01)

# valor_total_vendas = st.number_input("📈 Valor Total de Vendas:", min_value=0.0, step=0.01)
# numero_clientes = st.number_input("👥 Número de Clientes:", min_value=0, step=1)

# # Botão para gerar planilha e enviar
# if st.button("📅 Gerar Planilha e Enviar por Email"):
#     fechamento_data = {
#         "Data": [datetime.now().strftime("%Y-%m-%d")],
#         "Valor PIX (R$)": [valor_pix],
#         "Valor Dinheiro (R$)": [valor_dinheiro],
#         "Valor Cartão (R$)": [valor_cartao],
#         "Valor Pendura (R$)": [valor_pendura],
#         "Total de Entradas (R$)": [valor_pix + valor_dinheiro + valor_cartao + valor_pendura],
#         "Valor Total de Vendas (R$)": [valor_total_vendas],
#         "Número de Clientes": [numero_clientes],
#         "Divergência?": ["Sim" if (valor_pix + valor_dinheiro + valor_cartao + valor_pendura) != valor_total_vendas else "Não"]
#     }

#     df_fechamento = pd.DataFrame(fechamento_data)

#     nome_arquivo = f"fechamento_caixa_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
#     df_fechamento.to_excel(nome_arquivo, index=False)

#     try:
#         yag = yagmail.SMTP(user="ale.moreira@gmail.com", password="gncuqrzzkstgeamn")
#         yag.send(
#             to="ale.moreira@gmail.com",
#             subject="📋 Fechamento de Caixa - Villa Sonali",
#             contents="Segue em anexo o fechamento de caixa realizado hoje.",
#             attachments=nome_arquivo
#         )
#         st.success("Email enviado com sucesso!")
#         st.session_state.fechamento = {}
#     except Exception as e:
#         st.error(f"Erro ao enviar email: {e}")





# import streamlit as st
# import pandas as pd
# from datetime import datetime
# import yagmail

# st.title("📦 Fechamento de Caixa - Villa Sonali")

# # Entradas
# valor_pix = st.number_input("💳 Valor em PIX (R$):", min_value=0.0, step=0.01)
# valor_dinheiro = st.number_input("💵 Valor em Dinheiro (R$):", min_value=0.0, step=0.01)
# valor_cartao = st.number_input("💳 Valor em Cartão (R$):", min_value=0.0, step=0.01)
# valor_pendura = st.number_input("🧾 Valor Pendura (R$):", min_value=0.0, step=0.01)
# valor_total_vendas = st.number_input("💰 Valor Total de Vendas (com 10%) (R$):", min_value=0.0, step=0.01)
# numero_clientes = st.number_input("👥 Número de Clientes:", min_value=1, step=1)

# # Botão
# if st.button("📥 Gerar Planilha e Enviar por E-mail"):
#     total_entradas = round(valor_pix + valor_dinheiro + valor_cartao + valor_pendura, 2)
#     divergente = "✅ Sem divergência" if total_entradas == round(valor_total_vendas, 2) else "❌ Divergência detectada"
#     valor_bruto = round(valor_total_vendas / 1.10, 2)
#     ticket_medio = round(valor_total_vendas / numero_clientes, 2)

#     # Cria a planilha
#     df = pd.DataFrame({
#         "Tipo": [
#             "Valor PIX", "Valor Dinheiro", "Valor Cartão", "Valor Pendura", "", 
#             "Valor Total de Vendas (com 10%)", "Valor de Venda Bruto (sem 10%)", 
#             "Número de Clientes", "Ticket Médio", "Verificação"
#         ],
#         "Valor (R$)": [
#             valor_pix, valor_dinheiro, valor_cartao, valor_pendura, "", 
#             valor_total_vendas, valor_bruto, numero_clientes, ticket_medio, divergente
#         ]
#     })

#     # Salva o arquivo
#     data_hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
#     nome_arquivo = f"fechamento_caixa_{data_hora}.xlsx"
#     df.to_excel(nome_arquivo, index=False)

#     # Mostra mensagem de sucesso e botão de download
#     st.success(f"✅ Planilha gerada com sucesso: `{nome_arquivo}`")
#     with open(nome_arquivo, "rb") as file:
#         st.download_button("📂 Baixar Planilha", data=file, file_name=nome_arquivo,
#                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

#     # Envia por e-mail
#     try:
#         yag = yagmail.SMTP(user="ale.moreira@gmail.com", password="gncuqrzzkstgeamn")
#         yag.send(
#             to="ale.moreira@gmail.com",
#             subject="📋 Fechamento de Caixa - Villa Sonali",
#             contents="Segue em anexo o fechamento de caixa do dia.",
#             attachments=nome_arquivo
#         )
#         st.success("📧 E-mail enviado com sucesso!")
#     except Exception as e:
#         st.error(f"❌ Erro ao enviar e-mail: {e}")



# import streamlit as st
# import pandas as pd
# from datetime import datetime
# import yagmail
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill

# st.title("📦 Fechamento de Caixa - Villa Sonali")

# # Entradas do usuário
# valor_pix = st.number_input("💳 Valor em PIX (R$):", min_value=0.0, step=0.01)
# valor_dinheiro = st.number_input("💵 Valor em Dinheiro (R$):", min_value=0.0, step=0.01)
# valor_cartao = st.number_input("💳 Valor em Cartão (R$):", min_value=0.0, step=0.01)
# valor_pendura = st.number_input("🧾 Valor Pendura (R$):", min_value=0.0, step=0.01)
# valor_total_vendas = st.number_input("💰 Valor Total de Vendas (com 10%) (R$):", min_value=0.0, step=0.01)
# numero_clientes = st.number_input("👥 Número de Clientes:", min_value=1, step=1)

# # Botão para gerar e enviar planilha
# if st.button("📤 Gerar e Enviar Planilha por E-mail"):
#     total_entradas = round(valor_pix + valor_dinheiro + valor_cartao + valor_pendura, 2)
#     divergente = "✅ Sem divergência" if total_entradas == round(valor_total_vendas, 2) else "❌ Divergência detectada"
#     valor_bruto = round(valor_total_vendas / 1.10, 2)
#     ticket_medio = round(valor_total_vendas / numero_clientes, 2)

#     df = pd.DataFrame({
#         "Tipo": [
#             "Valor PIX",
#             "Valor Dinheiro",
#             "Valor Cartão",
#             "Valor Pendura",
#             "",
#             "Valor Total de Vendas (com 10%)",
#             "Valor de Venda Bruto (sem 10%)",
#             "Número de Clientes",
#             "Ticket Médio",
#             "Verificação"
#         ],
#         "Valor (R$)": [
#             valor_pix,
#             valor_dinheiro,
#             valor_cartao,
#             valor_pendura,
#             "",
#             valor_total_vendas,
#             valor_bruto,
#             numero_clientes,
#             ticket_medio,
#             divergente
#         ]
#     })

#     nome_arquivo = f"fechamento_caixa_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
#     df.to_excel(nome_arquivo, index=False)

#     # Formatação condicional no Excel
#     wb = load_workbook(nome_arquivo)
#     ws = wb.active

#     # Formatação da verificação (linha 11, coluna 2)
#     fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde
#     fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # vermelho

#     if divergente.startswith("✅"):
#         ws["B11"].fill = fill_verde
#     else:
#         ws["B11"].fill = fill_vermelho

#     # Formatação do ticket médio (linha 10, coluna 2)
#     if ticket_medio >= 100:
#         ws["B10"].fill = fill_verde
#     else:
#         ws["B10"].fill = fill_vermelho

#     wb.save(nome_arquivo)

#     # Envia o e-mail com yagmail
#     try:
#         yag = yagmail.SMTP(user="ale.moreira@gmail.com", password="gncuqrzzkstgeamn")
#         yag.send(
#             to="ale.moreira@gmail.com",
#             subject="📋 Relatório - Fechamento de Caixa",
#             contents="Segue em anexo o fechamento de caixa com os detalhes do dia.",
#             attachments=nome_arquivo
#         )
#         st.success("📧 E-mail enviado com sucesso!")
#     except Exception as e:
#         st.error(f"❌ Erro ao enviar o e-mail: {e}")


import streamlit as st
import pandas as pd
from datetime import datetime
import yagmail
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.title("📦 Fechamento de Caixa - Villa Sonali")

# Entradas do usuário
valor_pix = st.number_input("💳 Valor em PIX (R$):", min_value=0.0, step=0.01)
valor_dinheiro = st.number_input("💵 Valor em Dinheiro (R$):", min_value=0.0, step=0.01)
valor_cartao = st.number_input("💳 Valor em Cartão (R$):", min_value=0.0, step=0.01)
valor_pendura = st.number_input("🧾 Valor Pendura (R$):", min_value=0.0, step=0.01)
valor_caixa_inicial = st.number_input("💼 Dinheiro Inicial no Caixa (R$):", min_value=0.0, step=0.01)
valor_total_vendas = st.number_input("💰 Valor Total de Vendas (com 10%) (R$):", min_value=0.0, step=0.01)
numero_clientes = st.number_input("👥 Número de Clientes:", min_value=1, step=1)

# Botão para gerar e enviar planilha
if st.button("📤 Gerar e Enviar Planilha por E-mail"):
    total_entradas = round(valor_pix + valor_dinheiro + valor_cartao + valor_pendura, 2)
    divergencia_valor = round(total_entradas - valor_total_vendas, 2)
    valor_bruto = round(valor_total_vendas / 1.10, 2)
    ticket_medio = round(valor_total_vendas / numero_clientes, 2)

    if divergencia_valor == 0:
        divergente = "✅ Sem divergência"
    else:
        divergente = divergencia_valor

    df = pd.DataFrame({
        "Tipo": [
            "Valor PIX",
            "Valor Dinheiro",
            "Valor Cartão",
            "Valor Pendura",
            "Dinheiro Inicial no Caixa",
            "",
            "Valor Total de Vendas (com 10%)",
            "Valor de Venda Bruto (sem 10%)",
            "Número de Clientes",
            "Ticket Médio",
            "Verificação (diferença)"
        ],
        "Valor (R$)": [
            valor_pix,
            valor_dinheiro,
            valor_cartao,
            valor_pendura,
            valor_caixa_inicial,
            "",
            valor_total_vendas,
            valor_bruto,
            numero_clientes,
            ticket_medio,
            divergente
        ]
    })

    nome_arquivo = f"fechamento_caixa_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    df.to_excel(nome_arquivo, index=False)

    # Formatação condicional no Excel
    wb = load_workbook(nome_arquivo)
    ws = wb.active

    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde
    fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # vermelho

    # Formatação do ticket médio (linha 10, coluna 2)
    if ticket_medio >= 100:
        ws["B10"].fill = fill_verde
    else:
        ws["B10"].fill = fill_vermelho

    # Formatação da divergência (linha 11, coluna 2)
    if divergente == "✅ Sem divergência":
        ws["B11"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    elif isinstance(divergente, (float, int)):
        ws["B11"].fill = fill_vermelho if divergente < 0 else fill_verde

    wb.save(nome_arquivo)

    try:
        yag = yagmail.SMTP(user="ale.moreira@gmail.com", password="gncuqrzzkstgeamn")
        yag.send(
            to="ale.moreira@gmail.com",
            subject="📋 Relatório - Fechamento de Caixa",
            contents="Segue em anexo o fechamento de caixa com os detalhes do dia.",
            attachments=nome_arquivo
        )
        st.success("📧 E-mail enviado com sucesso!")
    except Exception as e:
        st.error(f"❌ Erro ao enviar o e-mail: {e}")





